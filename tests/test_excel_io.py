import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from lottery_draw_app.excel_io import load_prizes_from_excel, save_results_to_excel


class ExcelIoTests(unittest.TestCase):
    def test_load_prizes_reads_first_column(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "prizes.xlsx"
            pd.DataFrame({0: ["Gift 1", "Gift 2"], 1: ["ignore", "ignore"]}).to_excel(
                path, index=False, header=False
            )

            prizes = load_prizes_from_excel(path)
            self.assertEqual(prizes, ["Gift 1", "Gift 2"])

    def test_save_results_creates_two_columns(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "results.xlsx"
            save_results_to_excel(
                [{"Prize": "Gift 1", "Drawn Ticket": "Α01"}, {"Prize": "Gift 2", "Drawn Ticket": "Β11"}],
                path,
            )

            dataframe = pd.read_excel(path, engine="openpyxl")
            self.assertEqual(list(dataframe.columns), ["Δώρο", "Λαχνός"])
            self.assertEqual(len(dataframe), 2)

    def test_save_results_formatting(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "results.xlsx"
            save_results_to_excel(
                [{"Prize": "Gift 1", "Drawn Ticket": "Α01"}, {"Prize": "Gift 2", "Drawn Ticket": "Β11"}],
                path,
            )

            workbook = load_workbook(path)
            worksheet = workbook.active

            # Greek headers in row 1
            self.assertEqual(worksheet.cell(row=1, column=1).value, "Δώρο")
            self.assertEqual(worksheet.cell(row=1, column=2).value, "Λαχνός")

            # Headers are bold
            self.assertTrue(worksheet.cell(row=1, column=1).font.bold)
            self.assertTrue(worksheet.cell(row=1, column=2).font.bold)

            # Headers are center-aligned
            self.assertEqual(worksheet.cell(row=1, column=1).alignment.horizontal, "center")
            self.assertEqual(worksheet.cell(row=1, column=2).alignment.horizontal, "center")

            # Header background is RGB(191, 191, 191) → hex BFBFBF
            self.assertEqual(worksheet.cell(row=1, column=1).fill.fgColor.rgb[-6:].upper(), "BFBFBF")
            self.assertEqual(worksheet.cell(row=1, column=2).fill.fgColor.rgb[-6:].upper(), "BFBFBF")

            # Ticket column (col 2) data rows are center-aligned
            self.assertEqual(worksheet.cell(row=2, column=2).alignment.horizontal, "center")

            # First row is frozen
            self.assertEqual(str(worksheet.freeze_panes), "A2")

            # Column widths are at least max content + 7
            for col_idx in (1, 2):
                col_letter = get_column_letter(col_idx)
                max_len = max(
                    len(str(cell.value)) if cell.value is not None else 0
                    for cell in worksheet[col_letter]
                )
                self.assertAlmostEqual(
                    worksheet.column_dimensions[col_letter].width,
                    max_len + 7,
                )


if __name__ == "__main__":
    unittest.main()
