"""Excel import/export helpers."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


class ExcelDataError(ValueError):
    """Raised when data loaded from Excel is invalid."""


def load_prizes_from_excel(file_path: str | Path) -> list[str]:
    """Load one prize per row from the first column of an Excel file."""
    path = Path(file_path)
    if not path.exists():
        raise ExcelDataError(f"Το αρχείο δώρων δεν υπάρχει: {file_path}")

    try:
        dataframe = pd.read_excel(path, header=None, engine="openpyxl")
    except Exception as exc:
        raise ExcelDataError(f"Αποτυχία ανάγνωσης αρχείου Excel δώρων: {exc}") from exc

    if dataframe.empty:
        raise ExcelDataError("Το αρχείο Excel είναι κενό")

    first_column = dataframe.iloc[:, 0]
    prizes = [str(value).strip() for value in first_column if pd.notna(value) and str(value).strip()]

    if not prizes:
        raise ExcelDataError("Δεν βρέθηκαν έγκυρα δώρα στην πρώτη στήλη")

    return prizes


_GREEK_HEADERS = ("Δώρο", "Λαχνός")
_HEADER_FILL_COLOR = "BFBFBF"


def save_results_to_excel(results: list[dict[str, str]], file_path: str | Path) -> None:
    """Save draw results in exactly two columns with Greek headers and formatting."""
    if not results:
        raise ExcelDataError("Δεν υπάρχουν αποτελέσματα κλήρωσης για εξαγωγή")

    normalized_results = [
        {
            "Prize": row["Prize"],
            "Drawn Ticket": row["Drawn Ticket"],
        }
        for row in results
    ]

    dataframe = pd.DataFrame(normalized_results, columns=["Prize", "Drawn Ticket"])
    output_path = Path(file_path)

    try:
        dataframe.to_excel(output_path, index=False, header=False, engine="openpyxl")

        workbook = load_workbook(output_path)
        worksheet = workbook.active

        header_font = Font(bold=True)
        header_fill = PatternFill(fill_type="solid", fgColor=_HEADER_FILL_COLOR)
        center_align = Alignment(horizontal="center", vertical="center")

        worksheet.insert_rows(1)
        for col_idx, header_text in enumerate(_GREEK_HEADERS, start=1):
            cell = worksheet.cell(row=1, column=col_idx, value=header_text)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        ticket_col_idx = 2
        for row in worksheet.iter_rows(min_row=2, min_col=ticket_col_idx, max_col=ticket_col_idx):
            for cell in row:
                cell.alignment = center_align

        for col_idx in range(1, len(_GREEK_HEADERS) + 1):
            col_letter = get_column_letter(col_idx)
            max_length = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in worksheet[col_letter]
            )
            worksheet.column_dimensions[col_letter].width = max_length + 7

        worksheet.freeze_panes = "A2"

        workbook.save(output_path)
    except Exception as exc:
        raise ExcelDataError(f"Αποτυχία αποθήκευσης αποτελεσμάτων σε Excel: {exc}") from exc
